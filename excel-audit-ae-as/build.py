# -*- coding: utf-8 -*-
"""Reconstruction du classeur d'évaluation AE / AS d'Al Barid Bank (CRC).

- automatise l'onglet « Historique AE AS » (aucune saisie)
- corrige les anomalies A1 à A9 relevées dans AUDIT.md
- ajoute « Fiche Agent », « KPI Agents » et « Analyse Critères »

Entrée : SRC (classeur original, jamais modifié)   Sortie : DST
"""
from copy import copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl, coordinate_to_tuple
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.chart import LineChart, BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList

SRC = "/root/.claude/uploads/df2a0f54-c801-5124-a6bc-46534cbb9ac3/eedbb5a5-Grille_evaluation_abb_AE_AS__VCF.xlsx"
DST = "/home/user/foot-model/excel-audit-ae-as/Grille_evaluation_AE_AS_automatisee.xlsx"

# ---------------------------------------------------------------- palette (celle du fichier)
NAVY, NAVY2, BLUE = "1B365D", "2A4D7C", "2E5B9A"
ORANGE, LIGHT, GREY, BORDER, WHITE = "E87722", "E6EEF8", "555555", "BFCCE0", "FFFFFF"
YELLOW = "FFF2CC"
GREEN, GREEN_T = "C6EFCE", "006100"
RED, RED_T = "FFC7CE", "9C0006"
AMBER, AMBER_T = "FFEB9C", "9C6500"
TEAL = "3D8C8C"

FONT = "Calibri"
thin = Side(style="thin", color=BORDER)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

def f(sz=11, b=False, color="000000"):
    return Font(name=FONT, size=sz, bold=b, color=color)

def fill(c):
    return PatternFill("solid", fgColor=c)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
LEFTW = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

PCT, PCT2, NUM = "0.0%", "0.0%;[Red]-0.0%", "0"
MIN = '0.0" min"'
TREND = '▲ 0.0%;▼ 0.0%;"= "0.0%'

AE, AS_ = "Grille AE", "Grille AS"
HIST, KPI, CRIT = "Historique AE AS", "KPI Agents", "Analyse Critères"
HEB, EVO, SYN, FICHE = "Suivi Hebdo", "Evolution", "Synthèse & Performance", "Fiche Agent"
NEVAL = 60
EVCOLS = [gcl(7 + i) for i in range(NEVAL)]      # G .. BN

CIBLE = "'KPI Agents'!$E$5"          # cible de score global (80 %)
SEUIL = "'KPI Agents'!$F$5"          # complétude minimale pour qu'un audit compte

H = "'Historique AE AS'!"
def hc(col):
    return f"{H}${col}$4:${col}$123"
HA, HB, HC, HD, HE, HF, HG = (hc(c) for c in "ABCDEFG")
HH, HI, HJ, HK, HL, HM, HN = (hc(c) for c in "HIJKLMN")
HO, HP, HQ, HR, HS, HT = (hc(c) for c in "OPQRST")

wb = openpyxl.load_workbook(SRC)

# ================================================================ A9 — bandeau « A. ÉVALUATION DU FOND »
# La grille AE n'avait pas le bandeau de section que porte la grille AS : on insère
# la ligne manquante (l. 14) pour rendre les deux grilles structurellement identiques.
gae = wb[AE]
band_style = {k: copy(getattr(gae["B25"], k)) for k in ("font", "fill", "alignment")}
band_height = gae.row_dimensions[25].height
merges = [str(m) for m in gae.merged_cells.ranges]
for m in merges:
    gae.unmerge_cells(m)
heights = {r: gae.row_dimensions[r].height for r in range(14, gae.max_row + 1)
           if gae.row_dimensions[r].height}
gae.insert_rows(14)
for r in sorted(heights, reverse=True):
    gae.row_dimensions[r + 1].height = heights[r]
for m in merges:
    cr = CellRange(m)
    if cr.min_row >= 14:
        cr.shift(row_shift=1)
    elif cr.max_row >= 14:
        cr.max_row += 1
    gae.merge_cells(str(cr))
gae["B14"] = "A. ÉVALUATION DU FOND (Pondération : 70% dans le Score Global)"
for k, v in band_style.items():
    setattr(gae["B14"], k, v)
gae.merge_cells("B14:BN14")
gae.row_dimensions[14].height = band_height or 20
# A10 : le fichier d'origine figeait les volets en A38 sur la grille AE, soit 37 lignes
# bloquées (~1 300 px) — plus haut qu'un écran, donc feuille impossible à faire défiler.
# On aligne sur la grille AS : en-têtes figés jusqu'à la ligne des libellés de colonnes.
gae.freeze_panes = "A16"
for dv in list(gae.data_validations.dataValidation):
    if dv.type == "list" and dv.formula1 and "10,5,0" in str(dv.formula1):
        dv.sqref = openpyxl.worksheet.cell_range.MultiCellRange("G16:BN25 G27:BN41")

# lignes utiles des deux grilles, après insertion
GRID = {
    AE:  dict(name=3, log=4, date=5, start=6, end=7, ecoute=8, typo=9, objet=10,
              fond=(16, 25), forme=(27, 41), mfond=44, mforme=45, score=46, statut=47),
    AS_: dict(name=3, log=4, date=5, start=6, end=7, ecoute=8, typo=9, objet=10,
              fond=(16, 22), forme=(24, 31), mfond=34, mforme=35, score=36, statut=37),
}

# ================================================================ A1/A5 — correctifs des grilles
# A1 : la colonne E de la grille AS avait été figée à 10 avec une liste de saisie ;
#      la synthèse AS affichait donc 100 % en permanence.
# A5 : les lignes de synthèse renvoyaient #DIV/0! sur chaque colonne inutilisée.
for sheet in (AE, AS_):
    g, gs = GRID[sheet], wb[sheet]
    for lo, hi in (g["fond"], g["forme"]):
        for r in range(lo, hi + 1):
            c = gs[f"E{r}"]
            c.value = f'=IFERROR(AVERAGE(G{r}:BN{r}),"")'
            c.number_format = "0.00"
            c.fill = fill(LIGHT)
            c.alignment = CENTER
    for col in ["E"] + EVCOLS:
        fo, ff = g["fond"], g["forme"]
        gs[f'{col}{g["mfond"]}'] = (f'=IFERROR(AVERAGEIFS({col}{fo[0]}:{col}{fo[1]},'
                                    f'{col}{fo[0]}:{col}{fo[1]},"<>N/A")/10,"")')
        gs[f'{col}{g["mforme"]}'] = (f'=IFERROR(AVERAGEIFS({col}{ff[0]}:{col}{ff[1]},'
                                     f'{col}{ff[0]}:{col}{ff[1]},"<>N/A")/10,"")')
        gs[f'{col}{g["score"]}'] = (f'=IF(OR(NOT(ISNUMBER({col}{g["mfond"]})),'
                                    f'NOT(ISNUMBER({col}{g["mforme"]}))),"",'
                                    f'{col}{g["mfond"]}*0.7+{col}{g["mforme"]}*0.3)')
        gs[f'{col}{g["statut"]}'] = (f'=IF(ISNUMBER({col}{g["score"]}),'
                                     f'IF({col}{g["score"]}>={CIBLE},'
                                     '"Objectif Atteint ✅","Non Atteint ❌"),"")')

# A9 : la grille AS n'avait pas de ligne « STATUT OBJECTIF »
gas = wb[AS_]
gas["D37"] = "STATUT OBJECTIF :"
gas["D37"].font = f(11, True, NAVY); gas["D37"].alignment = RIGHT
for col in ["E"] + EVCOLS:
    gas[f"{col}37"].font = f(10, True); gas[f"{col}37"].alignment = CENTER
for dv in list(gas.data_validations.dataValidation):
    if dv.type == "list" and dv.formula1 and "10,5,0" in str(dv.formula1):
        dv.sqref = openpyxl.worksheet.cell_range.MultiCellRange("G16:BN22 G24:BN31")

# ================================================================ A2/A7/A8 — Historique automatisé
ws = wb[HIST]
for mr in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(mr))
ws.delete_rows(1, ws.max_row)

HEAD = ["Date d'écoute", "Nom de l'Agent", "LOG Agent", "Type d'appel", "Objet de l'appel",
        "DMC (min)", "Moyenne Fond", "Moyenne Forme", "Score Global", "Statut", "Semaine",
        "Année", "Source", "Colonne Audit", "Typologie de l'appel", "Type d'écoute",
        "Critères renseignés", "Complétude", "Audit exploitable", "Mois"]
WIDTHS = [13.5, 24, 11, 12, 16, 10, 13, 14, 12, 16, 9, 8, 11, 13, 22, 15, 17, 12, 16, 7]
NH = len(HEAD)
R_AE, R_AS = (4, 63), (64, 123)

ws["A1"] = "HISTORIQUE DES ÉVALUATIONS AE / AS"
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NH)
ws["A1"].font = f(16, True, WHITE); ws["A1"].fill = fill(NAVY); ws["A1"].alignment = CENTER
ws.row_dimensions[1].height = 30
ws["A2"] = ("Feuille 100 % calculée — aucune saisie ici. Chaque ligne reflète une colonne "
            "d'évaluation des onglets « Grille AE » et « Grille AS » ; elle se remplit dès qu'un "
            "nom d'agent est choisi dans la grille correspondante.")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NH)
ws["A2"].font = f(9, False, GREY); ws["A2"].alignment = LEFT
ws.row_dimensions[2].height = 26
for i, (h, w) in enumerate(zip(HEAD, WIDTHS), start=1):
    c = ws.cell(3, i, h)
    c.font = f(10, True, WHITE); c.fill = fill(NAVY2); c.alignment = CENTER; c.border = BOX
    ws.column_dimensions[gcl(i)].width = w
ws.row_dimensions[3].height = 34

def idx(sheet, row, i):
    return f"INDEX('{sheet}'!$G${row}:$BN${row},1,{i})"

def block(sheet, first, last, ncrit):
    g = GRID[sheet]
    for r in range(first, last + 1):
        i = r - first + 1
        nm, dt = idx(sheet, g["name"], i), idx(sheet, g["date"], i)
        st, en = idx(sheet, g["start"], i), idx(sheet, g["end"], i)
        obj, typ, eco = (idx(sheet, g[k], i) for k in ("objet", "typo", "ecoute"))
        gd = f'=IF($B{r}="","",'
        fo, ff = g["fond"], g["forme"]
        cnt = (f'COUNTA(INDEX(\'{sheet}\'!$G${fo[0]}:$BN${fo[1]},1,{i}):'
               f'INDEX(\'{sheet}\'!$G${fo[0]}:$BN${fo[1]},{fo[1]-fo[0]+1},{i}))'
               f'+COUNTA(INDEX(\'{sheet}\'!$G${ff[0]}:$BN${ff[1]},1,{i}):'
               f'INDEX(\'{sheet}\'!$G${ff[0]}:$BN${ff[1]},{ff[1]-ff[0]+1},{i}))')
        vals = {
            "A": f'=IF(OR({nm}="",{dt}=""),"",{dt})',
            "B": f'=IFERROR(IF({nm}="","",{nm}),"")',
            "C": f'{gd}IFERROR({idx(sheet,g["log"],i)},""))',
            "D": f'{gd}"{"AE" if sheet == AE else "AS"}")',
            "E": f'{gd}IF({obj}="","",{obj}))',
            "F": f'{gd}IF(OR({st}="",{en}=""),"",ROUND(MOD({en}-{st},1)*1440,0)))',
            "G": f'{gd}IFERROR({idx(sheet,g["mfond"],i)},""))',
            "H": f'{gd}IFERROR({idx(sheet,g["mforme"],i)},""))',
            "I": f'{gd}IFERROR({idx(sheet,g["score"],i)},""))',
            "J": f'=IF(NOT(ISNUMBER($I{r})),"",IF($I{r}>={CIBLE},"Objectif Atteint","Non Atteint"))',
            "K": f'=IF(NOT(ISNUMBER($A{r})),"",_xlfn.ISOWEEKNUM($A{r}))',
            "L": f'=IF(NOT(ISNUMBER($A{r})),"",YEAR($A{r}))',
            "M": f'{gd}"{sheet}")',
            "N": f'{gd}SUBSTITUTE(ADDRESS(1,6+{i},4),"1",""))',
            "O": f'{gd}IF({typ}="","",{typ}))',
            "P": f'{gd}IF({eco}="","",{eco}))',
            "Q": f'{gd}{cnt})',
            "R": f'=IF(NOT(ISNUMBER($Q{r})),"",$Q{r}/{ncrit})',
            "S": f'=IF(NOT(ISNUMBER($R{r})),"",IF($R{r}>={SEUIL},"Oui","Non"))',
            "T": f'=IF(NOT(ISNUMBER($A{r})),"",MONTH($A{r}))',
        }
        for col, v in vals.items():
            ws[f"{col}{r}"] = v
            c = ws[f"{col}{r}"]
            c.font = f(10); c.border = BOX
            c.alignment = LEFT if col in "BEMO" else CENTER
        ws[f"A{r}"].number_format = "dd/mm/yyyy"
        ws[f"F{r}"].number_format = MIN
        for col in "GHIR":
            ws[f"{col}{r}"].number_format = PCT
            ws[f"{col}{r}"].fill = fill(LIGHT)
        ws[f"I{r}"].font = f(10, True, NAVY)
        ws[f"T{r}"].number_format = NUM

block(AE, *R_AE, ncrit=25)
block(AS_, *R_AS, ncrit=15)

ws.freeze_panes = "C4"      # date + nom d'agent visibles en défilant vers la droite
ws.auto_filter.ref = f"A3:{gcl(NH)}{R_AS[1]}"
ws.sheet_view.showGridLines = False
for txt, bg, fg in (("Objectif Atteint", GREEN, GREEN_T), ("Non Atteint", RED, RED_T)):
    ws.conditional_formatting.add(f"J4:J{R_AS[1]}", FormulaRule(
        formula=[f'$J4="{txt}"'], fill=fill(bg), font=Font(name=FONT, size=10, color=fg)))
for txt, bg, fg in (("Oui", GREEN, GREEN_T), ("Non", AMBER, AMBER_T)):
    ws.conditional_formatting.add(f"S4:S{R_AS[1]}", FormulaRule(
        formula=[f'$S4="{txt}"'], fill=fill(bg), font=Font(name=FONT, size=10, bold=True, color=fg)))
ws.conditional_formatting.add(f"R4:R{R_AS[1]}", FormulaRule(
    formula=[f'AND(ISNUMBER($R4),$R4<{SEUIL})'],
    font=Font(name=FONT, size=10, color=AMBER_T)))

# ================================================================ KPI Agents
agents = [wb["LOG Agent"].cell(r, 1).value for r in range(2, 53)]
NAG = len(agents)
R0, R1 = 12, 12 + NAG - 1
TOT = R1 + 1

if KPI in wb.sheetnames:
    del wb[KPI]
k = wb.create_sheet(KPI, wb.sheetnames.index(HIST) + 1)

KH = ["Collaborateur", "Log", "Audits", "Exploitables", "dont AE", "dont AS",
      "Moyenne Fond", "Moyenne Forme", "Score global", "Écart vs cible",
      "Taux de conformité", "Complétude grille", "DMC moyenne", "Dernier score",
      "Score précédent", "Tendance", "Meilleur", "Moins bon", "Niveau", "Rang"]
KW = [26, 8, 8, 12, 8, 8, 11, 11, 12, 11, 12, 12, 11, 11, 12, 11, 10, 11, 16, 7]
NC, LAST = len(KH), gcl(len(KH))
for i, w in enumerate(KW, start=1):
    k.column_dimensions[gcl(i)].width = w

k.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NC)
k["A2"] = "BILAN & PERFORMANCE PAR AGENT"
k["A2"].font = f(15, True, WHITE); k["A2"].fill = fill(NAVY); k["A2"].alignment = CENTER
k.row_dimensions[2].height = 30
k.merge_cells(start_row=3, start_column=1, end_row=3, end_column=NC)
k["A3"] = ("AL BARID BANK — Centre de Relations Clients · Pondération 70 % Fond / 30 % Forme · "
           "toutes les valeurs sont calculées depuis l'onglet « Historique AE AS »")
k["A3"].font = f(9, False, ORANGE); k["A3"].alignment = CENTER

FL = [("Année", 2026, NUM), ("Type d'appel", "Tous", "General"), ("Semaine de", 1, NUM),
      ("Semaine à", 53, NUM), ("Cible score global", 0.8, PCT), ("Complétude minimale", 0.8, PCT)]
for i, (lab, val, fmt) in enumerate(FL, start=1):
    c = k.cell(4, i, lab); c.font = f(9, True, GREY); c.alignment = CENTER
    v = k.cell(5, i, val); v.font = f(12, True, NAVY); v.alignment = CENTER
    v.fill = fill(YELLOW); v.border = BOX; v.number_format = fmt
k.cell(4, 7, "◀ cellules jaunes : pilotent tout le tableau. « Cible » et « Complétude minimale » "
             "sont reprises par les grilles, l'historique et la fiche agent.")
k.cell(4, 7).font = f(9, False, GREY); k.cell(4, 7).alignment = LEFT
k.row_dimensions[5].height = 22
dv_type = DataValidation(type="list", formula1='"Tous,AE,AS"', allow_blank=False,
                         showDropDown=False, showErrorMessage=True, showInputMessage=True)
k.add_data_validation(dv_type); dv_type.add(k["B5"])

FILT = f'{HL},$A$5,{HK},">="&$C$5,{HK},"<="&$D$5,{HD},IF($B$5="Tous","*",$B$5)'

CARDS = [
    ("AUDITS RÉALISÉS", f'=COUNTIFS({FILT})', NUM, (1, 3)),
    ("DONT EXPLOITABLES", f'=COUNTIFS({FILT},{HS},"Oui")', NUM, (4, 6)),
    ("AGENTS AUDITÉS", f'=COUNTIF($C${R0}:$C${R1},">0")', NUM, (7, 8)),
    ("SCORE GLOBAL MOYEN", f'=IFERROR(AVERAGEIFS({HI},{FILT}),"")', PCT, (9, 10)),
    ("TAUX DE CONFORMITÉ", f'=IF($A$8=0,"",COUNTIFS({FILT},{HJ},"Objectif Atteint")/$A$8)', PCT, (11, 12)),
    ("MOYENNE FOND (70%)", f'=IFERROR(AVERAGEIFS({HG},{FILT}),"")', PCT, (13, 14)),
    ("MOYENNE FORME (30%)", f'=IFERROR(AVERAGEIFS({HH},{FILT}),"")', PCT, (15, 17)),
    ("COMPLÉTUDE MOYENNE", f'=IFERROR(AVERAGEIFS({HR},{FILT}),"")', PCT, (18, 20)),
]
for lab, formula, fmt, (c1, c2) in CARDS:
    k.merge_cells(start_row=7, start_column=c1, end_row=7, end_column=c2)
    k.merge_cells(start_row=8, start_column=c1, end_row=9, end_column=c2)
    hcell = k.cell(7, c1, lab)
    hcell.font = f(8, True, WHITE); hcell.fill = fill(BLUE); hcell.alignment = CENTER
    v = k.cell(8, c1, formula)
    v.font = f(18, True, NAVY); v.fill = fill(LIGHT); v.alignment = CENTER; v.number_format = fmt
    for rr in (7, 8, 9):
        for cc in range(c1, c2 + 1):
            k.cell(rr, cc).border = BOX
k.row_dimensions[7].height = 16
k.row_dimensions[9].height = 20

for i, h in enumerate(KH, start=1):
    c = k.cell(11, i, h)
    c.font = f(10, True, WHITE); c.fill = fill(NAVY2); c.alignment = CENTER; c.border = BOX
k.row_dimensions[11].height = 38

def kpi_row(r, name=None):
    per_agent = name is not None
    crit = f'{HB},$A{r},{FILT}' if per_agent else FILT
    last = f'_xlfn.MAXIFS({HA},{crit})'
    prev = f'_xlfn.MAXIFS({HA},{crit},{HA},"<"&{last})'
    vals = {
        "A": name if per_agent else "ENSEMBLE DU CENTRE (sélection)",
        "B": f'=IFERROR(VLOOKUP($A{r},\'LOG Agent\'!$A:$B,2,FALSE),"")' if per_agent else "",
        "C": f'=COUNTIFS({crit})',
        "D": f'=COUNTIFS({crit},{HS},"Oui")',
        "E": f'=COUNTIFS({crit},{HD},"AE")',
        "F": f'=COUNTIFS({crit},{HD},"AS")',
        "G": f'=IFERROR(AVERAGEIFS({HG},{crit}),"")',
        "H": f'=IFERROR(AVERAGEIFS({HH},{crit}),"")',
        "I": f'=IFERROR(AVERAGEIFS({HI},{crit}),"")',
        "J": f'=IF(NOT(ISNUMBER($I{r})),"",$I{r}-$E$5)',
        "K": f'=IF($C{r}=0,"",COUNTIFS({crit},{HJ},"Objectif Atteint")/$C{r})',
        "L": f'=IFERROR(AVERAGEIFS({HR},{crit}),"")',
        "M": f'=IFERROR(AVERAGEIFS({HF},{crit}),"")',
        "N": f'=IF($C{r}=0,"",IFERROR(AVERAGEIFS({HI},{crit},{HA},{last}),""))',
        "O": f'=IF($C{r}<2,"",IFERROR(AVERAGEIFS({HI},{crit},{HA},{prev}),""))',
        "P": f'=IF(OR(NOT(ISNUMBER($N{r})),NOT(ISNUMBER($O{r}))),"",$N{r}-$O{r})',
        "Q": f'=IF($C{r}=0,"",_xlfn.MAXIFS({HI},{crit}))',
        "R": f'=IF($C{r}=0,"",_xlfn.MINIFS({HI},{crit}))',
        "S": (f'=IF(NOT(ISNUMBER($I{r})),"",IF($I{r}>=0.9,"Excellent",'
              f'IF($I{r}>=$E$5,"Conforme",IF($I{r}>=$E$5-0.1,"À accompagner","Critique"))))'),
        "T": f'=IF(NOT(ISNUMBER($I{r})),"",RANK($I{r},$I${R0}:$I${R1}))' if per_agent else "",
    }
    for col, v in vals.items():
        c = k[f"{col}{r}"]
        c.value = v if v != "" else None
        c.font = f(10) if per_agent else f(10, True, NAVY)
        c.border = BOX
        c.alignment = LEFT if col == "A" else CENTER
    for col in "GHIJKLNOPQR":
        k[f"{col}{r}"].number_format = PCT
    k[f"J{r}"].number_format = PCT2
    k[f"P{r}"].number_format = TREND
    k[f"M{r}"].number_format = MIN
    for col in "CDEFT":
        k[f"{col}{r}"].number_format = NUM
    k[f"I{r}"].font = f(11, True, NAVY)
    k[f"I{r}"].fill = fill(LIGHT)
    if not per_agent:
        for col in KH and [gcl(i) for i in range(1, NC + 1)]:
            k[f"{col}{r}"].fill = fill(LIGHT)

for j, nm in enumerate(agents):
    kpi_row(R0 + j, nm)
kpi_row(TOT)

k.freeze_panes = "C12"
k.auto_filter.ref = f"A11:{LAST}{R1}"
k.sheet_view.showGridLines = False
k.conditional_formatting.add(f"I{R0}:I{R1}", ColorScaleRule(
    start_type="num", start_value=0.6, start_color="F8696B",
    mid_type="num", mid_value=0.8, mid_color="FFEB84",
    end_type="num", end_value=1.0, end_color="63BE7B"))
k.conditional_formatting.add(f"K{R0}:K{R1}", DataBarRule(
    start_type="num", start_value=0, end_type="num", end_value=1, color="2E5B9A"))
k.conditional_formatting.add(f"L{R0}:L{R1}", DataBarRule(
    start_type="num", start_value=0, end_type="num", end_value=1, color=ORANGE))
for lvl, bg, fg in (("Excellent", GREEN, GREEN_T), ("Conforme", "DDEBF7", "1F4E79"),
                    ("À accompagner", AMBER, AMBER_T), ("Critique", RED, RED_T)):
    k.conditional_formatting.add(f"S{R0}:S{R1}", FormulaRule(
        formula=[f'$S{R0}="{lvl}"'], fill=fill(bg),
        font=Font(name=FONT, size=10, bold=True, color=fg)))
k.conditional_formatting.add(f"A{R0}:{LAST}{R1}", FormulaRule(
    formula=[f'$C{R0}=0'], font=Font(name=FONT, size=10, color="B7B7B7")))
k.conditional_formatting.add(f"P{R0}:P{R1}", FormulaRule(
    formula=[f'AND(ISNUMBER($P{R0}),$P{R0}<0)'],
    font=Font(name=FONT, size=10, bold=True, color=RED_T)))
k.conditional_formatting.add(f"P{R0}:P{R1}", FormulaRule(
    formula=[f'AND(ISNUMBER($P{R0}),$P{R0}>0)'],
    font=Font(name=FONT, size=10, bold=True, color=GREEN_T)))
k.conditional_formatting.add(f"D{R0}:D{R1}", FormulaRule(
    formula=[f'AND($C{R0}>0,$D{R0}<$C{R0})'],
    fill=fill(AMBER), font=Font(name=FONT, size=10, color=AMBER_T)))

for i, txt in enumerate((
        "Lecture : « Exploitables » = audits dont la complétude atteint le seuil de la cellule F5. "
        "Un score élevé sur une grille peu remplie n'est pas représentatif.",
        "Niveaux : Excellent ≥ 90 % · Conforme ≥ cible · À accompagner ≥ cible − 10 pts · Critique en dessous.")):
    rr = TOT + 2 + i
    k.cell(rr, 1, txt).font = f(9, False, GREY)
    k.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=NC)

# ================================================================ Analyse Critères
if CRIT in wb.sheetnames:
    del wb[CRIT]
a = wb.create_sheet(CRIT, wb.sheetnames.index(KPI) + 1)
for i, w in enumerate([7, 11, 10, 78, 15, 11, 14, 18], start=1):
    a.column_dimensions[gcl(i)].width = w
a.merge_cells("A2:H2")
a["A2"] = "ANALYSE PAR CRITÈRE — OÙ SE JOUENT LES POINTS PERDUS"
a["A2"].font = f(15, True, WHITE); a["A2"].fill = fill(NAVY); a["A2"].alignment = CENTER
a.row_dimensions[2].height = 30
a.merge_cells("A3:H3")
a["A3"] = ("Moyenne de chaque critère sur l'ensemble des évaluations saisies dans la grille "
           "correspondante (colonnes G à BN). Les critères notés « N/A » sont neutralisés.")
a["A3"].font = f(9, False, ORANGE); a["A3"].alignment = CENTER
for i, h in enumerate(["Réf", "Grille", "Catégorie", "Critère d'évaluation", "Note moyenne /10",
                       "Score", "Nb évaluations", "Statut"], start=1):
    c = a.cell(5, i, h)
    c.font = f(10, True, WHITE); c.fill = fill(NAVY2); c.alignment = CENTER; c.border = BOX
a.row_dimensions[5].height = 30

CRIT_ROWS = []          # (grille, catégorie, ligne de la grille) — réutilisé par la fiche agent
r = 6
first_data = r
for sheet, label in ((AE, "Grille AE"), (AS_, "Grille AS")):
    g = GRID[sheet]
    a.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = a.cell(r, 1, f"{label} — {'appels entrants' if sheet == AE else 'appels sortants'}")
    c.font = f(11, True, WHITE); c.fill = fill(BLUE); c.alignment = CENTER
    r += 1
    for cat, (lo, hi) in (("Fond", g["fond"]), ("Forme", g["forme"])):
        for gr in range(lo, hi + 1):
            CRIT_ROWS.append((sheet, cat, gr))
            a.cell(r, 1, f"='{sheet}'!$B{gr}")
            a.cell(r, 2, label)
            a.cell(r, 3, cat)
            a.cell(r, 4, f"='{sheet}'!$D{gr}")
            a.cell(r, 5, f'=IFERROR(AVERAGE(\'{sheet}\'!$G{gr}:$BN{gr}),"")')
            a.cell(r, 6, f'=IF(NOT(ISNUMBER($E{r})),"",$E{r}/10)')
            a.cell(r, 7, f"=COUNT('{sheet}'!$G{gr}:$BN{gr})")
            a.cell(r, 8, f'=IF(NOT(ISNUMBER($F{r})),"—",'
                         f'IF($F{r}>=0.9,"Maîtrisé",IF($F{r}>=0.7,"À consolider","Point critique")))')
            for cc in range(1, 9):
                cell = a.cell(r, cc)
                cell.font = f(10); cell.border = BOX
                cell.alignment = LEFTW if cc == 4 else CENTER
            a.cell(r, 5).number_format = "0.00"
            a.cell(r, 6).number_format = PCT
            a.cell(r, 6).fill = fill(LIGHT); a.cell(r, 6).font = f(10, True, NAVY)
            a.cell(r, 7).number_format = NUM
            a.row_dimensions[r].height = 28
            r += 1
last_data = r - 1
a.freeze_panes = "A6"
a.sheet_view.showGridLines = False
a.conditional_formatting.add(f"F{first_data}:F{last_data}", DataBarRule(
    start_type="num", start_value=0, end_type="num", end_value=1, color="2E5B9A"))
for lvl, bg, fg in (("Maîtrisé", GREEN, GREEN_T), ("À consolider", AMBER, AMBER_T),
                    ("Point critique", RED, RED_T)):
    a.conditional_formatting.add(f"H{first_data}:H{last_data}", FormulaRule(
        formula=[f'$H{first_data}="{lvl}"'], fill=fill(bg),
        font=Font(name=FONT, size=10, bold=True, color=fg)))

# ================================================================ Suivi Hebdo (dynamique)
s = wb[HEB]
for mr in list(s.merged_cells.ranges):
    s.unmerge_cells(str(mr))
s.delete_rows(1, s.max_row)
for i, w in enumerate([11, 14, 14, 14, 14, 15, 14, 14], start=1):
    s.column_dimensions[gcl(i)].width = w
s.merge_cells("A1:H1")
s["A1"] = "SUIVI HEBDOMADAIRE DES PERFORMANCES"
s["A1"].font = f(15, True, WHITE); s["A1"].fill = fill(NAVY); s["A1"].alignment = CENTER
s.row_dimensions[1].height = 30
s["A3"], s["A4"] = "Année", "Agent suivi"
for cc in ("A3", "A4"):
    s[cc].font = f(10, True, GREY); s[cc].alignment = RIGHT
s["B3"], s["B4"] = 2026, agents[1]
for cc in ("B3", "B4"):
    s[cc].font = f(11, True, NAVY); s[cc].fill = fill(YELLOW)
    s[cc].border = BOX; s[cc].alignment = CENTER
s.merge_cells("B4:C4")
s["D3"] = "◀ cellules de saisie : pilotent le tableau et les graphiques de l'onglet « Evolution »."
s["D3"].font = f(9, False, GREY); s["D3"].alignment = LEFT
dv_ag = DataValidation(type="list", formula1="'LOG Agent'!$A$2:$A$52", allow_blank=True,
                       showDropDown=False, showErrorMessage=True, showInputMessage=True)
s.add_data_validation(dv_ag); dv_ag.add(s["B4"])
for i, h in enumerate(["Semaine", "Audits centre", "Score centre", "Fond centre", "Forme centre",
                       "Taux conformité", "Audits agent", "Score agent"], start=1):
    c = s.cell(6, i, h)
    c.font = f(10, True, WHITE); c.fill = fill(NAVY2); c.alignment = CENTER; c.border = BOX
s.row_dimensions[6].height = 30
HEB_R0, HEB_R1 = 7, 59
for w in range(1, 54):
    r = HEB_R0 + w - 1
    base = f'{HL},$B$3,{HK},$A{r}'
    s.cell(r, 1, w).number_format = '"S"0'
    s.cell(r, 2, f'=COUNTIFS({base})')
    s.cell(r, 3, f'=IFERROR(AVERAGEIFS({HI},{base}),"")')
    s.cell(r, 4, f'=IFERROR(AVERAGEIFS({HG},{base}),"")')
    s.cell(r, 5, f'=IFERROR(AVERAGEIFS({HH},{base}),"")')
    s.cell(r, 6, f'=IF($B{r}=0,"",COUNTIFS({base},{HJ},"Objectif Atteint")/$B{r})')
    s.cell(r, 7, f'=COUNTIFS({base},{HB},$B$4)')
    s.cell(r, 8, f'=IFERROR(AVERAGEIFS({HI},{base},{HB},$B$4),"")')
    for cc in range(1, 9):
        cell = s.cell(r, cc); cell.font = f(10); cell.border = BOX; cell.alignment = CENTER
    for cc in (3, 4, 5, 6, 8):
        s.cell(r, cc).number_format = PCT
    for cc in (2, 7):
        s.cell(r, cc).number_format = NUM
    s.cell(r, 3).fill = fill(LIGHT); s.cell(r, 8).fill = fill(LIGHT)
s.freeze_panes = "A7"
s.sheet_view.showGridLines = False
s.conditional_formatting.add(f"C{HEB_R0}:C{HEB_R1}", ColorScaleRule(
    start_type="num", start_value=0.6, start_color="F8696B",
    mid_type="num", mid_value=0.8, mid_color="FFEB84",
    end_type="num", end_value=1.0, end_color="63BE7B"))
s.conditional_formatting.add(f"B{HEB_R0}:B{HEB_R1}", DataBarRule(
    start_type="num", start_value=0, end_type="max", color="9DC3E6"))

# ================================================================ Evolution (graphiques centre)
e = wb[EVO]
e.delete_rows(1, max(e.max_row, 3))
for ch in list(getattr(e, "_charts", [])):
    e._charts.remove(ch)
e.merge_cells("A1:P1")
e["A1"] = "ÉVOLUTION DES PERFORMANCES AE / AS"
e["A1"].font = f(15, True, WHITE); e["A1"].fill = fill(NAVY); e["A1"].alignment = CENTER
e.row_dimensions[1].height = 30
e.merge_cells("A2:P2")
e["A2"] = ("Graphiques alimentés par l'onglet « Suivi Hebdo ». Changez l'année ou l'agent suivi "
           "dans les cellules jaunes de cet onglet : les courbes se recalculent.")
e["A2"].font = f(9, False, ORANGE); e["A2"].alignment = CENTER
e.sheet_view.showGridLines = False
for i in range(1, 17):
    e.column_dimensions[gcl(i)].width = 9
cats = Reference(s, min_col=1, min_row=HEB_R0, max_row=HEB_R1)

def line_chart(sheet, anchor, title, cols, colors, src, cat_ref, r0, r1, ytitle="Score", pct=True):
    ch = LineChart()
    ch.title, ch.style = title, 2
    ch.height, ch.width = 8.5, 17
    ch.y_axis.title, ch.x_axis.title = ytitle, "Semaine ISO"
    ch.dispBlanksAs = "gap"
    for col in cols:
        ch.add_data(Reference(src, min_col=col, min_row=r0 - 1, max_row=r1), titles_from_data=True)
    ch.set_categories(cat_ref)
    for ser, colr in zip(ch.series, colors):
        ser.graphicalProperties.line.solidFill = colr
        ser.graphicalProperties.line.width = 22000
        ser.smooth = False
    if pct:
        ch.y_axis.numFmt = "0%"
    sheet.add_chart(ch, anchor)

line_chart(e, "A4", "Score global — centre vs agent suivi", [3, 8], [NAVY, ORANGE], s, cats, HEB_R0, HEB_R1)
line_chart(e, "I4", "Fond (70 %) vs Forme (30 %) — centre", [4, 5], [BLUE, "70AD47"], s, cats, HEB_R0, HEB_R1)
line_chart(e, "A22", "Taux de conformité hebdomadaire (part des audits ≥ cible)", [6], [ORANGE], s, cats, HEB_R0, HEB_R1)
bar = BarChart(); bar.type = "col"; bar.style = 2
bar.title = "Volume d'audits réalisés par semaine"
bar.height, bar.width = 8.5, 17
bar.y_axis.title, bar.x_axis.title = "Nb d'audits", "Semaine ISO"
bar.add_data(Reference(s, min_col=2, min_row=HEB_R0 - 1, max_row=HEB_R1), titles_from_data=True)
bar.set_categories(cats)
bar.series[0].graphicalProperties.solidFill = BLUE
e.add_chart(bar, "I22")

# ================================================================ Fiche Agent (tableau de bord individuel)
ot = wb["Objet&TYP"]
objets, seen = [], set()
for col, lo, hi in (("F", 8, 16), ("A", 8, 18)):
    for rr in range(lo, hi + 1):
        v = ot[f"{col}{rr}"].value
        if v and str(v).strip() and str(v).strip() not in seen:
            seen.add(str(v).strip())
            objets.append(str(v).strip())

if FICHE in wb.sheetnames:
    del wb[FICHE]
fa = wb.create_sheet(FICHE, wb.sheetnames.index(SYN) + 1)
NCOL = 20
for i in range(1, NCOL + 1):
    fa.column_dimensions[gcl(i)].width = 9
fa.sheet_view.showGridLines = False
fa.freeze_panes = "A9"      # le nom de l'agent reste visible en faisant défiler les graphiques

AGT = "$C$8"
FA = f'{HB},{AGT},{HL},$F$6'          # audits de l'agent, année sélectionnée
FC = f'{HL},$F$6'                     # audits du centre, même année

def merge_set(rng, value, font_, fill_=None, align=CENTER, fmt=None):
    fa.merge_cells(rng)
    c = fa[rng.split(":")[0]]
    c.value = value; c.font = font_; c.alignment = align
    if fill_:
        c.fill = fill(fill_)
    if fmt:
        c.number_format = fmt
    return c

merge_set(f"A2:{gcl(NCOL)}2", "FICHE AGENT — TABLEAU DE BORD INDIVIDUEL", f(15, True, WHITE), NAVY)
fa.row_dimensions[2].height = 30
merge_set(f"A3:{gcl(NCOL)}3",
          "Saisissez un nom d'agent OU un log HERMES : tous les indicateurs et graphiques ci-dessous "
          "se recalculent automatiquement. Le log est prioritaire sur le nom s'il est renseigné.",
          f(9, False, ORANGE))

for rng, lab in (("A5:C5", "Nom de l'agent"), ("D5:E5", "ou Log HERMES"), ("F5:G5", "Année")):
    merge_set(rng, lab, f(9, True, GREY))
for rng, val, fmt in (("A6:C6", agents[1], "General"), ("D6:E6", None, NUM), ("F6:G6", 2026, NUM)):
    c = merge_set(rng, val, f(12, True, NAVY), YELLOW, fmt=fmt)
    c.border = BOX
fa.row_dimensions[6].height = 22
dv_fiche = DataValidation(type="list", formula1="'LOG Agent'!$A$2:$A$52", allow_blank=True,
                          showDropDown=False, showErrorMessage=True, showInputMessage=True)
fa.add_data_validation(dv_fiche); dv_fiche.add(fa["A6"])
dv_log = DataValidation(type="list", formula1="'LOG Agent'!$B$2:$B$52", allow_blank=True,
                        showDropDown=False, showErrorMessage=True, showInputMessage=True)
fa.add_data_validation(dv_log); dv_log.add(fa["D6"])

merge_set("A8:B8", "AGENT ANALYSÉ →", f(9, True, GREY), align=RIGHT)
merge_set("C8:J8",
          '=IFERROR(IF($D$6<>"",INDEX(\'LOG Agent\'!$A$2:$A$52,'
          'MATCH($D$6,\'LOG Agent\'!$B$2:$B$52,0)),$A$6),$A$6)',
          f(16, True, NAVY), LIGHT)
merge_set("K8:M8", f'=IF({AGT}="","","Log : "&IFERROR(VLOOKUP({AGT},\'LOG Agent\'!$A:$B,2,FALSE),"?"))',
          f(11, True, GREY))
merge_set("N8:P8",
          f'=IF({AGT}="","","Rang : "&IFERROR(TEXT(INDEX(\'KPI Agents\'!$T${R0}:$T${R1},'
          f'MATCH({AGT},\'KPI Agents\'!$A${R0}:$A${R1},0)),"0")&" / "&'
          f'TEXT(COUNTIF(\'KPI Agents\'!$C${R0}:$C${R1},">0"),"0"),"n/a"))',
          f(11, True, NAVY))
merge_set("Q8:T8",
          f'=IF({AGT}="","⚠ Saisissez un nom ou un log.",'
          f'IF(COUNTIFS({FA})=0,"⚠ Aucun audit sur l\'année sélectionnée.",""))',
          f(10, True, AMBER_T))
fa.row_dimensions[8].height = 26

# ---- bloc de données (sous les graphiques) --------------------------------
D0 = 92
merge_set(f"A{D0}:{gcl(NCOL)}{D0}",
          "DONNÉES DES GRAPHIQUES — calculées automatiquement, ne rien modifier",
          f(11, True, WHITE), NAVY2)

def data_header(row, titles):
    for i, t in enumerate(titles, start=1):
        c = fa.cell(row, i, t)
        c.font = f(9, True, WHITE); c.fill = fill(BLUE); c.alignment = CENTER; c.border = BOX

def data_row(row, label, formulas, fmt=PCT):
    c = fa.cell(row, 1, label); c.font = f(9); c.alignment = LEFT; c.border = BOX
    for i, formula in enumerate(formulas, start=2):
        cc = fa.cell(row, i, formula)
        cc.font = f(9); cc.alignment = CENTER; cc.border = BOX; cc.number_format = fmt

# 1. indicateurs clés agent vs centre
IND_H = D0 + 2
data_header(IND_H, ["Indicateur", "Agent", "Centre"])
IND0 = IND_H + 1
IND = [("Moyenne Fond", HG), ("Moyenne Forme", HH), ("Score global", HI), (None, None), ("Complétude", HR)]
for j, (lab, rng) in enumerate(IND):
    rr = IND0 + j
    if lab is None:
        data_row(rr, "Taux de conformité",
                 [f'=IF(COUNTIFS({FA})=0,"",COUNTIFS({FA},{HJ},"Objectif Atteint")/COUNTIFS({FA}))',
                  f'=IF(COUNTIFS({FC})=0,"",COUNTIFS({FC},{HJ},"Objectif Atteint")/COUNTIFS({FC}))'])
    else:
        data_row(rr, lab, [f'=IFERROR(AVERAGEIFS({rng},{FA}),"")',
                           f'=IFERROR(AVERAGEIFS({rng},{FC}),"")'])
IND1 = IND0 + len(IND) - 1
C_FOND, C_FORME, C_SCORE, C_CONF, C_COMPL = (f"$B${IND0 + i}" for i in range(5))

# 2. évolution hebdomadaire
WK_H = IND1 + 2
data_header(WK_H, ["Semaine", "Agent", "Centre", "Cible"])
WK0 = WK_H + 1
for w in range(1, 54):
    rr = WK0 + w - 1
    fa.cell(rr, 1, w).number_format = '"S"0'
    data_row(rr, None, [f'=IFERROR(AVERAGEIFS({HI},{FA},{HK},$A{rr}),"")',
                        f'=IFERROR(AVERAGEIFS({HI},{FC},{HK},$A{rr}),"")',
                        f'={CIBLE}'])
    fa.cell(rr, 1, w).font = f(9); fa.cell(rr, 1).alignment = CENTER; fa.cell(rr, 1).border = BOX
WK1 = WK0 + 52

# 3. évolution mensuelle
MOIS = ["Janv.", "Févr.", "Mars", "Avril", "Mai", "Juin",
        "Juil.", "Août", "Sept.", "Oct.", "Nov.", "Déc."]
MO_H = WK1 + 2
data_header(MO_H, ["Mois", "Agent", "Centre"])
MO0 = MO_H + 1
for m, lab in enumerate(MOIS, start=1):
    rr = MO0 + m - 1
    data_row(rr, lab, [f'=IFERROR(AVERAGEIFS({HI},{FA},{HT},{m}),"")',
                       f'=IFERROR(AVERAGEIFS({HI},{FC},{HT},{m}),"")'])
MO1 = MO0 + 11

# 4. score par objet d'appel
OB_H = MO1 + 2
data_header(OB_H, ["Objet de l'appel", "Score agent", "Nb audits"])
OB0 = OB_H + 1
for j, obj in enumerate(objets):
    rr = OB0 + j
    data_row(rr, obj, [f'=IFERROR(AVERAGEIFS({HI},{FA},{HE},"{obj}*"),"")',
                       f'=COUNTIFS({FA},{HE},"{obj}*")'])
    fa.cell(rr, 3).number_format = NUM
OB1 = OB0 + len(objets) - 1

# 5. profil par critère (Fond puis Forme), toutes grilles confondues
def crit_block(header_row, cat):
    data_header(header_row, ["Critère", "Score agent", "Libellé complet"])
    rr = header_row + 1
    for sheet, c, gr in CRIT_ROWS:
        if c != cat:
            continue
        tag = "AE" if sheet == AE else "AS"
        fa.cell(rr, 1, f'="{tag} "&\'{sheet}\'!$B{gr}')
        fa.cell(rr, 2, f'=IFERROR(AVERAGEIF(\'{sheet}\'!$G$3:$BN$3,{AGT},'
                       f'\'{sheet}\'!$G{gr}:$BN{gr})/10,"")')
        fa.cell(rr, 3, f"='{sheet}'!$D{gr}")
        for cc in range(1, 4):
            cell = fa.cell(rr, cc)
            cell.font = f(9); cell.border = BOX
            cell.alignment = LEFT if cc == 3 else CENTER
        fa.cell(rr, 2).number_format = PCT
        rr += 1
    return header_row + 1, rr - 1

CF_H = OB1 + 2
CF0, CF1 = crit_block(CF_H, "Fond")
CR_H = CF1 + 2
CR0, CR1 = crit_block(CR_H, "Forme")

# 6. répartitions
TY_H = CR1 + 2
data_header(TY_H, ["Type d'appel", "Nb audits"])
TY0 = TY_H + 1
for j, t in enumerate(("AE", "AS")):
    data_row(TY0 + j, f"Appels {'entrants' if t == 'AE' else 'sortants'} ({t})",
             [f'=COUNTIFS({FA},{HD},"{t}")'], fmt=NUM)
TY1 = TY0 + 1

ST_H = TY1 + 2
data_header(ST_H, ["Statut", "Nb audits"])
ST0 = ST_H + 1
for j, t in enumerate(("Objectif Atteint", "Non Atteint")):
    data_row(ST0 + j, t, [f'=COUNTIFS({FA},{HJ},"{t}")'], fmt=NUM)
ST1 = ST0 + 1

# ---- cartes KPI -----------------------------------------------------------
FCARDS = [
    ("AUDITS", f'=COUNTIFS({FA})', NUM),
    ("EXPLOITABLES", f'=COUNTIFS({FA},{HS},"Oui")', NUM),
    ("SCORE GLOBAL", f'={C_SCORE}', PCT),
    ("ÉCART VS CIBLE", f'=IF(NOT(ISNUMBER({C_SCORE})),"",{C_SCORE}-{CIBLE})', PCT2),
    ("CONFORMITÉ", f'={C_CONF}', PCT),
    ("FOND (70%)", f'={C_FOND}', PCT),
    ("FORME (30%)", f'={C_FORME}', PCT),
    ("COMPLÉTUDE", f'={C_COMPL}', PCT),
    ("DMC MOYENNE", f'=IFERROR(AVERAGEIFS({HF},{FA}),"")', MIN),
    ("NIVEAU", f'=IF(NOT(ISNUMBER({C_SCORE})),"—",IF({C_SCORE}>=0.9,"Excellent",'
               f'IF({C_SCORE}>={CIBLE},"Conforme",IF({C_SCORE}>={CIBLE}-0.1,'
               '"À accompagner","Critique"))))', "General"),
]
for j, (lab, formula, fmt) in enumerate(FCARDS):
    c1 = 1 + 2 * j
    c2 = c1 + 1
    fa.merge_cells(start_row=10, start_column=c1, end_row=10, end_column=c2)
    fa.merge_cells(start_row=11, start_column=c1, end_row=12, end_column=c2)
    hcell = fa.cell(10, c1, lab)
    hcell.font = f(8, True, WHITE); hcell.fill = fill(BLUE); hcell.alignment = CENTER
    v = fa.cell(11, c1, formula)
    v.font = f(14 if lab != "NIVEAU" else 12, True, NAVY)
    v.fill = fill(LIGHT); v.alignment = CENTER; v.number_format = fmt
    for rr in (10, 11, 12):
        for cc in range(c1, c2 + 1):
            fa.cell(rr, cc).border = BOX
fa.row_dimensions[10].height = 16
fa.row_dimensions[12].height = 18

# ---- graphiques -----------------------------------------------------------
def bar_chart(anchor, title, h_row, r0, r1, ncols, colors, horizontal=False,
              height=8.5, width=16.5, fmt="0%", xtitle=None):
    ch = BarChart()
    ch.type = "bar" if horizontal else "col"
    ch.style, ch.title = 2, title
    ch.height, ch.width = height, width
    ch.gapWidth = 60
    for i in range(2, 2 + ncols):
        ch.add_data(Reference(fa, min_col=i, min_row=h_row, max_row=r1), titles_from_data=True)
    ch.set_categories(Reference(fa, min_col=1, min_row=r0, max_row=r1))
    for ser, colr in zip(ch.series, colors):
        ser.graphicalProperties.solidFill = colr
        ser.graphicalProperties.line.solidFill = colr
    if fmt:
        ch.y_axis.numFmt = fmt
    if xtitle:
        ch.x_axis.title = xtitle
    if ncols == 1:
        ch.legend = None
    fa.add_chart(ch, anchor)

def donut(anchor, title, h_row, r0, r1, colors):
    d = DoughnutChart(holeSize=52)
    d.title = title
    d.height, d.width = 8.5, 16.5
    d.add_data(Reference(fa, min_col=2, min_row=h_row, max_row=r1), titles_from_data=True)
    d.set_categories(Reference(fa, min_col=1, min_row=r0, max_row=r1))
    pts = d.series[0]
    from openpyxl.chart.marker import DataPoint
    pts.data_points = []
    for i, colr in enumerate(colors):
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill = colr
        pts.data_points.append(dp)
    d.dataLabels = DataLabelList()
    d.dataLabels.showVal = True
    fa.add_chart(d, anchor)

bar_chart("A14", "Indicateurs clés — agent vs centre", IND_H, IND0, IND1, 2, [NAVY, "BFCCE0"])
line_chart(fa, "K14", "Évolution hebdomadaire du score", [2, 3, 4], [ORANGE, NAVY, "C00000"],
           fa, Reference(fa, min_col=1, min_row=WK0, max_row=WK1), WK0, WK1)
bar_chart("A31", "Score mensuel — agent vs centre", MO_H, MO0, MO1, 2, [ORANGE, "BFCCE0"], xtitle="Mois")
bar_chart("K31", "Score moyen par objet d'appel", OB_H, OB0, OB1, 1, [TEAL])
bar_chart("A48", "Profil par critère — FOND (70 %)", CF_H, CF0, CF1, 1, [NAVY],
          horizontal=True, height=13)
bar_chart("K48", "Profil par critère — FORME (30 %)", CR_H, CR0, CR1, 1, [BLUE],
          horizontal=True, height=13)
donut("A75", "Répartition des audits (AE / AS)", TY_H, TY0, TY1, [NAVY, ORANGE])
donut("K75", "Objectif atteint / non atteint", ST_H, ST0, ST1, ["63BE7B", "F8696B"])

merge_set(f"A{D0 - 2}:{gcl(NCOL)}{D0 - 2}",
          "Les profils par critère portent sur toutes les écoutes de l'agent saisies dans les grilles "
          "(sans filtre d'année) ; les autres indicateurs suivent l'année sélectionnée.",
          f(9, False, GREY), align=LEFT)

# ================================================================ A3 — Synthèse & Performance
y = wb[SYN]
y["B9"] = f'=COUNT({HI})'
y["C9"] = f'=IFERROR(AVERAGE({HI}),"")'
y["C9"].number_format = PCT
y["D9"], y["E9"] = "=C13", "=F13"
y["F8"] = "TAUX DE CONFORMITÉ"
y["F9"] = f'=IF($B$9=0,"",COUNTIF({HJ},"Objectif Atteint")/$B$9)'
y["F9"].number_format = PCT
y["G9"] = "=_xlfn.ISOWEEKNUM(TODAY())"
for cell, rng, typ in (("C13", HI, "AE"), ("C14", HG, "AE"), ("C15", HH, "AE"),
                       ("F13", HI, "AS"), ("F14", HG, "AS"), ("F15", HH, "AS")):
    y[cell] = f'=IFERROR(AVERAGEIFS({rng},{HD},"{typ}"),"")'
for cell, src in (("C16", "C13"), ("F16", "F13")):
    y[cell] = (f'=IF(NOT(ISNUMBER({src})),"—",IF({src}>={CIBLE},'
               '"Objectif Atteint ✅","Non Atteint ❌"))')
y["D13"] = f'="Cible : ≥ "&TEXT({CIBLE},"0.0%")'
y["G13"] = f'="Cible : ≥ "&TEXT({CIBLE},"0.0%")'
y.merge_cells("B18:G18")
y["B18"] = ("Moyennes issues de l'onglet « Historique AE AS », lui-même alimenté automatiquement "
            "par les grilles AE et AS. Fiche individuelle : onglet « Fiche Agent » · "
            "tableau complet : « KPI Agents » · détail par critère : « Analyse Critères ».")
y["B18"].font = f(9, False, GREY)
y["B18"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
y.row_dimensions[18].height = 30

# ================================================================ finitions
for name, col in {SYN: NAVY, FICHE: ORANGE, AE: BLUE, AS_: BLUE, HIST: "70AD47",
                  KPI: ORANGE, CRIT: ORANGE, HEB: "7F7F7F", EVO: "7F7F7F"}.items():
    wb[name].sheet_properties.tabColor = col
for name, titles in ((KPI, "1:11"), (CRIT, "1:5"), (HIST, "1:3"), (HEB, "1:6"), (FICHE, "1:8")):
    p = wb[name]
    p.page_setup.orientation = "landscape"
    p.page_setup.fitToWidth = 1
    p.page_setup.fitToHeight = 0
    p.sheet_properties.pageSetUpPr.fitToPage = True
    p.print_title_rows = titles
# garde-fou : un volet figé plus grand qu'un écran rend la feuille non défilable (anomalie A10)
MAX_PT, MAX_CH = 320.0, 45.0
for w in wb.worksheets:
    if not w.freeze_panes:
        continue
    fr, fcol = coordinate_to_tuple(w.freeze_panes)
    ht = sum((w.row_dimensions[i].height or 15.0) for i in range(1, fr))
    wd = sum((w.column_dimensions[gcl(i)].width or 8.43) for i in range(1, fcol))
    assert ht <= MAX_PT and wd <= MAX_CH, (
        f"{w.title} : volet figé trop grand ({ht:.0f} pt / {wd:.1f} car.) — "
        f"la feuille ne pourrait pas défiler")

wb.active = wb.sheetnames.index(SYN)

wb.save(DST)
print("saved", DST)
print("onglets :", wb.sheetnames)
